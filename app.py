import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from datetime import datetime, timedelta
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Page configuration
st.set_page_config(
    page_title="Calculadora de Investimentos",
    page_icon="💰",
    layout="wide"
)

# Custom CSS
st.markdown("""
    <style>
    .main {
        padding: 2rem;
    }
    .stButton>button {
        width: 100%;
        background-color: #4CAF50;
        color: white;
        padding: 0.5rem 1rem;
        border-radius: 5px;
        border: none;
        font-weight: bold;
    }
    .stButton>button:hover {
        background-color: #45a049;
    }
    </style>
""", unsafe_allow_html=True)

# Header
st.title("💰 Calculadora de Investimentos")

# Short explanation using st.info
st.info(
    """
    **O que faz esta aplicação?**  
    Esta calculadora permite simular o crescimento do seu investimento ao longo do tempo, considerando juros compostos e reforços periódicos, até atingir uma meta financeira definida por si. Visualize o progresso período a período, veja projeções gráficas e analise os detalhes do seu plano de investimento.
    """
)

st.markdown("### Planeie o seu futuro financeiro com a nossa calculadora de juros compostos")

# Tutorial Section
with st.expander("📚 Como utilizar esta calculadora"):
    st.markdown("""
    ### Tutorial Rápido
    1. **Defina o seu objetivo**:
       - Introduza o valor que pretende alcançar (Meta)
       - Este é o seu objetivo financeiro final
    2. **Configure os seus investimentos**:
       - Saldo Inicial: Quanto já tem investido
       - Taxa de Juro: A taxa de retorno anual esperada (utilize o cursor)
       - Periodicidade: Escolha a frequência dos seus investimentos
       - Reforço: Quanto planeia reforçar em cada período
    3. **Visualize os resultados**:
       - O gráfico mostra o crescimento do seu investimento
       - A linha vermelha representa a sua meta
       - Clique em 'Mostrar Tabela' para ver os detalhes período a período
    ### Dicas importantes
    - Considere uma taxa de juros realista (geralmente entre 5% e 10% para investimentos de longo prazo)
    - O reforço periódico deve ser um valor que consegue manter consistentemente
    - Quanto mais cedo começar, mais tempo o juro composto tem para trabalhar a seu favor
    """)

# Sidebar for inputs
st.sidebar.markdown("## 📊 Introduza os seus dados")
alvo = st.sidebar.number_input("🎯 Valor que pretende alcançar (€)", min_value=0, step=1000)
Saldo_Inicial = st.sidebar.number_input("💰 Quanto já tem investido (€)", min_value=0, step=1000)

# Periodicity selection
st.sidebar.markdown("#### ⏱️ Periodicidade")
periodicidade = st.sidebar.selectbox(
    "Frequência dos investimentos",
    options=["Mensal", "Trimestral", "Semestral", "Anual"],
    help="Escolha com que frequência fará os seus investimentos"
)

# Map periodicity to months and periods per year
periodicidade_info = {
    "Mensal": {"meses": 1, "periodos_ano": 12, "adverbio": "mensalmente", "plural": "meses"},
    "Trimestral": {"meses": 3, "periodos_ano": 4, "adverbio": "trimestralmente", "plural": "trimestres"},
    "Semestral": {"meses": 6, "periodos_ano": 2, "adverbio": "semestralmente", "plural": "semestres"},
    "Anual": {"meses": 12, "periodos_ano": 1, "adverbio": "anualmente", "plural": "anos"}
}

# Interest rate input based on periodicity
st.sidebar.markdown("#### 📈 Taxa de Juro")
taxa_tipo = st.sidebar.radio(
    "Tipo de taxa",
    ["Taxa Anual", f"Taxa {periodicidade.lower()}"],
    help="Escolha se a taxa é anual ou por período"
)

# Initialize session state for interest rate
if 'taxa_juro' not in st.session_state:
    st.session_state.taxa_juro = 5.0

if taxa_tipo == "Taxa Anual":
    # Create two columns for the slider and number input
    col1, col2 = st.sidebar.columns([3, 1])
    with col1:
        Taxa_Juro = st.slider(
            "Taxa de juro anual (%)",
            min_value=0.0,
            max_value=35.0,
            value=st.session_state.taxa_juro,
            step=0.1,
            key="slider_taxa_anual",
            help="Deslize para ajustar a taxa de juro anual esperada",
            on_change=lambda: setattr(st.session_state, 'taxa_juro', st.session_state.slider_taxa_anual)
        )
    with col2:
        Taxa_Juro = st.number_input(
            "Taxa (%)",
            min_value=0.0,
            max_value=35.0,
            value=st.session_state.taxa_juro,
            step=0.1,
            key="input_taxa_anual",
            help="Introduza a taxa de juro anual com precisão",
            on_change=lambda: setattr(st.session_state, 'taxa_juro', st.session_state.input_taxa_anual)
        )
    # Update session state
    st.session_state.taxa_juro = Taxa_Juro
    # Convert annual rate to periodic rate
    periodos_ano = periodicidade_info[periodicidade]["periodos_ano"]
    taxa_periodo = ((1 + Taxa_Juro/100) ** (1/periodos_ano) - 1) * 100
else:
    # Create two columns for the slider and number input
    col1, col2 = st.sidebar.columns([3, 1])
    with col1:
        Taxa_Juro = st.slider(
            f"Taxa de juro {periodicidade.lower()} (%)",
            min_value=0.0,
            max_value=35.0,
            value=st.session_state.taxa_juro,
            step=0.1,
            key="slider_taxa_periodo",
            help=f"Deslize para ajustar a taxa de juro {periodicidade.lower()}",
            on_change=lambda: setattr(st.session_state, 'taxa_juro', st.session_state.slider_taxa_periodo)
        )
    with col2:
        Taxa_Juro = st.number_input(
            "Taxa (%)",
            min_value=0.0,
            max_value=35.0,
            value=st.session_state.taxa_juro,
            step=0.1,
            key="input_taxa_periodo",
            help=f"Introduza a taxa de juro {periodicidade.lower()} com precisão",
            on_change=lambda: setattr(st.session_state, 'taxa_juro', st.session_state.input_taxa_periodo)
        )
    # Update session state
    st.session_state.taxa_juro = Taxa_Juro
    # Convert periodic rate to annual rate
    periodos_ano = periodicidade_info[periodicidade]["periodos_ano"]
    taxa_periodo = Taxa_Juro
    Taxa_Juro = ((1 + Taxa_Juro/100) ** periodos_ano - 1) * 100

# Adjust reinforcement input based on periodicity
st.sidebar.markdown("#### ➕ Reforço")
Reforco = st.sidebar.number_input(
    f"Quanto vai reforçar {periodicidade_info[periodicidade]['adverbio']}? (€)",
    min_value=0,
    step=1000
)

calcular = st.sidebar.button("🔄 Calcular Projeção", key="calcular_btn")

# Calculation logic
if 'df_resultado' not in st.session_state:
    st.session_state['df_resultado'] = None
    st.session_state['alvo'] = None

if calcular:
    # Initialize lists for data collection
    periodos, datas, saldos_iniciais, taxas, juros_list, reforcos, saldos_finais = [], [], [], [], [], [], []
    
    # Initial values
    saldo = Saldo_Inicial
    periodo = 1
    data_atual = datetime.now()
    meses_por_periodo = periodicidade_info[periodicidade]["meses"]
    
    # Calculate periods until target is reached or 100 years
    while saldo < alvo and periodo <= (100 * 12 / meses_por_periodo):
        # Calculate interest for the period using the periodic rate
        juros = saldo * (taxa_periodo / 100)
        saldo_final = saldo + juros + Reforco
        
        # Store data
        periodos.append(periodo)
        datas.append(data_atual.strftime("%d/%m/%Y"))
        saldos_iniciais.append(saldo)
        taxas.append(f"{taxa_periodo:.2f}%")
        juros_list.append(juros)
        reforcos.append(Reforco)
        saldos_finais.append(saldo_final)
        
        # Update for next period
        saldo = saldo_final
        periodo += 1
        data_atual = data_atual + timedelta(days=30 * meses_por_periodo)
    
    # Create DataFrame
    df = pd.DataFrame({
        "Período": periodos,
        "Data": datas,
        "Saldo Inicial": saldos_iniciais,
        "Taxa de Juro": taxas,
        "Juros": juros_list,
        "Reforço": reforcos,
        "Saldo Final": saldos_finais
    })
    
    st.session_state["df_resultado"] = df
    st.session_state["alvo"] = alvo

# Main area for outputs only
if st.session_state["df_resultado"] is not None:
    df = st.session_state["df_resultado"]
    alvo = st.session_state["alvo"]
    
    # Summary metrics
    mcol1, mcol2, mcol3 = st.columns(3)
    with mcol1:
        st.metric(f"Períodos até à meta", f"{len(df)} {periodicidade_info[periodicidade]['plural']}")
    with mcol2:
        st.metric("Investimento total", f"€{df['Reforço'].sum() + Saldo_Inicial:,.2f}")
    with mcol3:
        st.metric("Juros acumulados", f"€{df['Juros'].sum():,.2f}")
    
    # Gráfico
    fig = go.Figure()
    
    # Add main investment line
    fig.add_trace(go.Scatter(
        x=df["Período"],
        y=df["Saldo Final"],
        mode="lines+markers",
        name="Saldo Final (€)",
        line=dict(color="#2E86C1", width=3),
        marker=dict(
            symbol="circle",
            size=8,
            color="#2E86C1",
            line=dict(color="white", width=2)
        ),
        hovertemplate="<b>Período %{x}</b><br>" +
                     "Saldo: €%{y:,.2f}<br>" +
                     "<extra></extra>"
    ))
    
    # Add interest gained line
    fig.add_trace(go.Scatter(
        x=df["Período"],
        y=df["Juros"],
        mode="lines+markers",
        name="Juros Ganhos (€)",
        line=dict(color="#27AE60", width=2, dash="dot"),
        marker=dict(
            symbol="diamond",
            size=6,
            color="#27AE60",
            line=dict(color="white", width=1)
        ),
        hovertemplate="<b>Período %{x}</b><br>" +
                     "Juros: €%{y:,.2f}<br>" +
                     "<extra></extra>"
    ))
    
    # Add reinforcements line
    fig.add_trace(go.Scatter(
        x=df["Período"],
        y=df["Reforço"],
        mode="lines+markers",
        name="Reforços (€)",
        line=dict(color="#E67E22", width=2, dash="dot"),
        marker=dict(
            symbol="square",
            size=6,
            color="#E67E22",
            line=dict(color="white", width=1)
        ),
        hovertemplate="<b>Período %{x}</b><br>" +
                     "Reforço: €%{y:,.2f}<br>" +
                     "<extra></extra>"
    ))
    
    # Add target line
    fig.add_trace(go.Scatter(
        x=[df["Período"].min(), df["Período"].max()],
        y=[alvo, alvo],
        mode="lines",
        name="Meta (€)",
        line=dict(color="#E74C3C", dash="dash", width=2),
        hovertemplate="<b>Meta</b><br>" +
                     "Valor: €%{y:,.2f}<br>" +
                     "<extra></extra>"
    ))
    
    # Add area under the curve
    fig.add_trace(go.Scatter(
        x=df["Período"],
        y=df["Saldo Final"],
        fill='tozeroy',
        fillcolor='rgba(46, 134, 193, 0.1)',
        line=dict(width=0),
        showlegend=False,
        hoverinfo='skip'
    ))
    
    # Update layout
    fig.update_layout(
        title=dict(
            text=f"Crescimento do Investimento ao Longo do Tempo ({periodicidade})",
            font=dict(size=24, color="#2C3E50"),
            x=0.5,
            y=0.95,
            xanchor="center"
        ),
        xaxis=dict(
            title=dict(
                text=f"Período ({periodicidade})",
                font=dict(size=14, color="#2C3E50", family="Arial"),
                standoff=10
            ),
            gridcolor='rgba(0,0,0,0.1)',
            showgrid=True,
            zeroline=True,
            zerolinecolor='rgba(0,0,0,0.2)',
            zerolinewidth=1,
            tickfont=dict(size=12, color="#2C3E50", family="Arial"),
            tickmode='linear',
            tick0=1,
            dtick=1,
            showline=True,
            linewidth=1,
            linecolor='rgba(0,0,0,0.2)',
            mirror=True
        ),
        yaxis=dict(
            title=dict(
                text="Valor (€)",
                font=dict(size=14, color="#2C3E50", family="Arial"),
                standoff=10
            ),
            gridcolor='rgba(0,0,0,0.1)',
            showgrid=True,
            zeroline=True,
            zerolinecolor='rgba(0,0,0,0.2)',
            zerolinewidth=1,
            tickfont=dict(size=12, color="#2C3E50", family="Arial"),
            tickformat=",.0f",
            showline=True,
            linewidth=1,
            linecolor='rgba(0,0,0,0.2)',
            mirror=True,
            separatethousands=True,
            exponentformat='none'
        ),
        legend=dict(
            yanchor="top",
            y=0.99,
            xanchor="left",
            x=0.01,
            bgcolor="rgba(255,255,255,0.8)",
            bordercolor="rgba(0,0,0,0.1)",
            borderwidth=1,
            font=dict(size=12, family="Arial")
        ),
        template="plotly_white",
        hovermode="x unified",
        plot_bgcolor='white',
        paper_bgcolor='white',
        margin=dict(t=100, b=50, l=80, r=50),
        height=600
    )
    
    # Add annotations for key points
    fig.add_annotation(
        x=df["Período"].iloc[-1],
        y=df["Saldo Final"].iloc[-1],
        text=f"Saldo Final: €{df['Saldo Final'].iloc[-1]:,.2f}",
        showarrow=True,
        arrowhead=2,
        arrowsize=1,
        arrowwidth=2,
        arrowcolor="#2E86C1",
        ax=50,
        ay=-50,
        font=dict(size=12, color="#2C3E50")
    )
    
    fig.add_annotation(
        x=0,
        y=alvo,
        text=f"Meta: €{alvo:,.2f}",
        showarrow=True,
        arrowhead=2,
        arrowsize=1,
        arrowwidth=2,
        arrowcolor="#E74C3C",
        ax=50,
        ay=50,
        font=dict(size=12, color="#2C3E50")
    )
    
    # Add range slider with improved styling
    fig.update_layout(
        xaxis=dict(
            rangeslider=dict(
                visible=True,
                thickness=0.1,
                bgcolor='rgba(0,0,0,0.05)',
                bordercolor='rgba(0,0,0,0.1)',
                borderwidth=1
            ),
            type="linear"
        )
    )
    
    st.plotly_chart(fig, use_container_width=True)
    
    # Detailed results with download functionality
    st.markdown("### 📋 Detalhes do Investimento")
    
    # Create Excel download function
    def create_excel_download(df, periodicidade):
        # Create a new workbook and select the active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Projeção de Investimento"
        
        # Add title
        ws['A1'] = f"Projeção de Investimento - {periodicidade}"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:G1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Add headers
        headers = ["Período", "Data", "Saldo Inicial (€)", "Taxa de Juro", "Juros (€)", "Reforço (€)", "Saldo Final (€)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 4):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx in [3, 5, 6, 7]:  # Currency columns
                    cell.number_format = '#,##0.00€'
                cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            # Use header row cell (row 3, index 2) to safely get column letter,
            # avoiding issues with merged cells in row 1.
            if not column or len(column) < 3:
                continue
            
            column_letter = column[2].column_letter
            for cell in column:
                if cell.value:
                    # Find the max length of the string representation of a cell's value
                    max_length = max(max_length, len(str(cell.value)))

            # Add padding and set column width
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer
    
    # Create download button
    if st.button("📊 Mostrar Tabela Detalhada", key="tabela_btn"):
        st.dataframe(
            df.style.format({
                "Saldo Inicial": "€{:.2f}",
                "Juros": "€{:.2f}",
                "Reforço": "€{:.2f}",
                "Saldo Final": "€{:.2f}"
            }).background_gradient(subset=["Saldo Final"], cmap="YlOrRd"),
            use_container_width=True
        )
        
        # Create and offer Excel download
        excel_buffer = create_excel_download(df, periodicidade)
        st.download_button(
            label="📥 Transferir Tabela em Excel",
            data=excel_buffer.getvalue(),
            file_name=f"projecao_investimento_{periodicidade.lower()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Clique para transferir a tabela detalhada em formato Excel"
        )

# Footer
st.markdown("---")
st.markdown("""
    <div style='text-align: center'>
        <p>💡 Lembre-se: Esta é uma projeção e os resultados reais podem variar.</p>
        <p>Consulte um profissional financeiro para orientação personalizada.</p>
    </div>
""", unsafe_allow_html=True)